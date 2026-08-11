"""Geração de planilhas Excel — formatação padrão compartilhada (cabeçalho
em negrito e congelado, linhas com altura 21) e os geradores específicos de
cada auditoria."""
import io

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

FILL_DENTRO = PatternFill("solid", fgColor="C6EFCE")     # verde
FILL_FORA = PatternFill("solid", fgColor="FFC7CE")       # vermelho
FILL_ABERTO = PatternFill("solid", fgColor="FFEB9C")     # âmbar
FILL_DUPLICADO = PatternFill("solid", fgColor="D9D9D9")  # cinza
FILL_CABECALHO = PatternFill("solid", fgColor="0C447C")  # azul escuro

FILL_POR_STATUS = {
    "Dentro do foco": FILL_DENTRO,
    "Fora do foco": FILL_FORA,
    "Duplicado": FILL_DUPLICADO,
}


def estilizar_cabecalho(ws):
    """Cabeçalho azul-escuro com a letra branca em negrito — padrão visual
    das planilhas geradas."""
    for cel in ws[1]:
        cel.fill = FILL_CABECALHO
        cel.font = Font(color="FFFFFF", bold=True)


def formatar_planilha_padrao(ws, altura_linha=21):
    """Cabeçalho congelado e todas as linhas com altura fixa. Mantém a COR da
    fonte do cabeçalho (branca) — antes reescrevia como Font(bold=True), que
    voltava pro preto e apagava a letra branca."""
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True, color=cell.font.color)
    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = altura_linha


def ajustar_largura_colunas(ws, largura_min=10, largura_max=70):
    """Ajusta cada coluna ao maior conteúdo (cabeçalho ou dado). Usa uma folga
    um pouco maior porque o cabeçalho vai em negrito, que ocupa mais espaço."""
    for col in ws.columns:
        comprimentos = [len(str(c.value)) for c in col if c.value is not None]
        if not comprimentos:
            continue
        larg = min(largura_max, max(largura_min, max(comprimentos) + 3))
        ws.column_dimensions[col[0].column_letter].width = larg


def gerar_xlsx_volume(resultado_bruto_df: pd.DataFrame) -> bytes:
    """Excel da auditoria de volume: aba "Resumo por Produto" (quantos não
    recebidos por produto, igual já existia) + uma única aba "Não
    recebidos" com todas as linhas, cada uma colorida pelo status de foco
    (verde/vermelho/âmbar) quando a IA classificou — mesmo padrão visual do
    Excel do validador de foco, sem dividir por abas de status."""
    wb = Workbook()

    ws_resumo = wb.active
    ws_resumo.title = "Resumo por Produto"
    if not resultado_bruto_df.empty and "Produto Consultado" in resultado_bruto_df.columns:
        resumo = (
            resultado_bruto_df.groupby("Produto Consultado")
            .size()
            .reset_index(name="Qtde Não Recebidos")
            .sort_values("Qtde Não Recebidos", ascending=False)
        )
    else:
        resumo = pd.DataFrame(columns=["Produto Consultado", "Qtde Não Recebidos"])
    ws_resumo.append(list(resumo.columns))
    for _, linha in resumo.iterrows():
        ws_resumo.append(list(linha))
    estilizar_cabecalho(ws_resumo)
    ajustar_largura_colunas(ws_resumo)
    formatar_planilha_padrao(ws_resumo)

    ws = wb.create_sheet("Não recebidos")
    colunas = list(resultado_bruto_df.columns)
    ws.append(colunas)
    estilizar_cabecalho(ws)
    tem_status = "STATUS_IA" in colunas
    for _, linha in resultado_bruto_df.iterrows():
        ws.append([linha[c] for c in colunas])
        if tem_status:
            fill = FILL_POR_STATUS.get(linha["STATUS_IA"], FILL_ABERTO)
            for cel in ws[ws.max_row]:
                cel.fill = fill
    ajustar_largura_colunas(ws)
    formatar_planilha_padrao(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def gerar_xlsx_validado(cabecalho, registros, leads, classificacoes) -> bytes:
    """Excel do validador de foco: uma aba com todas as linhas + STATUS +
    MOTIVO, pintando cada linha por status: verde (dentro), vermelho (fora),
    âmbar (aberto), cinza (duplicado — não passou pela IA)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads validados"
    colunas = cabecalho + ["STATUS", "MOTIVO"]
    ws.append(colunas)
    estilizar_cabecalho(ws)
    for i, r in enumerate(registros):
        c = classificacoes.get(leads[i]["id"], {
            "status": "Aberto", "motivo": "Não classificado pela IA — revisar manualmente.",
        })
        linha = list(r) + [c["status"], c["motivo"]]
        ws.append(linha)
        fill = FILL_POR_STATUS.get(c["status"], FILL_ABERTO)
        for cel in ws[ws.max_row]:
            cel.fill = fill
    ajustar_largura_colunas(ws)
    formatar_planilha_padrao(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def combinar_workbooks(pares: list) -> bytes:
    """Junta várias planilhas (cada uma já pronta em bytes) num único
    arquivo, uma aba por planilha de origem — usado pelo download
    'Excel combinado' da tela de saúde do cliente. pares: lista de
    (prefixo_nome_aba, xlsx_bytes)."""
    destino = Workbook()
    destino.remove(destino.active)
    nomes_usados = set()
    for prefixo, xlsx_bytes in pares:
        if not xlsx_bytes:
            continue
        origem = load_workbook(io.BytesIO(xlsx_bytes))
        for nome_aba in origem.sheetnames:
            ws_origem = origem[nome_aba]
            nome_final = f"{prefixo} - {nome_aba}"[:31]
            sufixo = 1
            base = nome_final
            while nome_final in nomes_usados:
                sufixo += 1
                nome_final = f"{base[:28]}~{sufixo}"
            nomes_usados.add(nome_final)
            ws_destino = destino.create_sheet(title=nome_final)
            for linha in ws_origem.iter_rows():
                for celula in linha:
                    nova = ws_destino.cell(row=celula.row, column=celula.column, value=celula.value)
                    if celula.has_style:
                        nova.font = celula.font.copy()
                        nova.fill = celula.fill.copy()
                        nova.alignment = celula.alignment.copy()
            ws_destino.freeze_panes = "A2"
            for row_idx in range(1, ws_destino.max_row + 1):
                ws_destino.row_dimensions[row_idx].height = 21
    buf = io.BytesIO()
    destino.save(buf)
    return buf.getvalue()
