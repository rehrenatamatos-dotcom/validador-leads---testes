"""Geração de planilhas Excel — formatação padrão compartilhada (cabeçalho
em negrito e congelado, linhas com altura 21) e os geradores específicos de
cada auditoria."""
import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

FILL_DENTRO = PatternFill("solid", fgColor="C6EFCE")     # verde
FILL_FORA = PatternFill("solid", fgColor="FFC7CE")       # vermelho
FILL_ABERTO = PatternFill("solid", fgColor="FFEB9C")     # âmbar
FILL_DUPLICADO = PatternFill("solid", fgColor="D9D9D9")  # cinza
FILL_CABECALHO = PatternFill("solid", fgColor="0C447C")  # azul escuro

FILL_POR_STATUS = {
    "Dentro do foco": FILL_DENTRO,
    "Fora do foco": FILL_FORA,
    "Duplicado": FILL_DUPLICADO,
}


def formatar_planilha_padrao(ws, altura_linha=21):
    """Cabeçalho em negrito e congelado, todas as linhas com altura fixa —
    formatação padrão das duas ferramentas originais."""
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = altura_linha


def ajustar_largura_colunas(ws, largura_min=12, largura_max=60):
    for col in ws.columns:
        larg = min(largura_max, max(largura_min, max(
            (len(str(c.value)) if c.value else 0) for c in col
        ) + 2))
        ws.column_dimensions[col[0].column_letter].width = larg


def gerar_xlsx_volume(resultado_perdido_df, descartados_df=None, abertos_df=None) -> bytes:
    """Excel da auditoria de volume: Resumo + Perdidos (dentro do foco) e,
    se houver filtro de foco aplicado, também Fora de foco (descartado) e
    Aberto (incerto) — nada some silenciosamente do arquivo."""
    buffer = io.BytesIO()
    import pandas as pd
    from openpyxl.styles import Font as _Font

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        if not resultado_perdido_df.empty and "Produto Consultado" in resultado_perdido_df.columns:
            resumo = (
                resultado_perdido_df.groupby("Produto Consultado")
                .size()
                .reset_index(name="Qtde Perdidos")
                .sort_values("Qtde Perdidos", ascending=False)
            )
        else:
            resumo = pd.DataFrame(columns=["Produto Consultado", "Qtde Perdidos"])
        resumo.to_excel(writer, sheet_name="Resumo por Produto", index=False)
        resultado_perdido_df.to_excel(writer, sheet_name="Perdidos", index=False)
        if descartados_df is not None and not descartados_df.empty:
            descartados_df.to_excel(writer, sheet_name="Descartados (fora de foco)", index=False)
        if abertos_df is not None and not abertos_df.empty:
            abertos_df.to_excel(writer, sheet_name="Aberto (incerto)", index=False)

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.font = _Font(bold=True)
            for row_idx in range(1, ws.max_row + 1):
                ws.row_dimensions[row_idx].height = 21

    return buffer.getvalue()


def gerar_xlsx_validado(cabecalho, registros, leads, classificacoes) -> bytes:
    """Excel do validador de foco: uma aba com todas as linhas + STATUS +
    MOTIVO, pintando cada linha por status: verde (dentro), vermelho (fora),
    âmbar (aberto), cinza (duplicado — não passou pela IA)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads validados"
    colunas = cabecalho + ["STATUS", "MOTIVO"]
    ws.append(colunas)
    for cel in ws[1]:
        cel.fill = FILL_CABECALHO
        cel.font = Font(color="FFFFFF", bold=True)
    for i, r in enumerate(registros):
        c = classificacoes.get(leads[i]["id"], {
            "status": "Aberto", "motivo": "Não classificado pela IA — revisar manualmente.",
        })
        linha = list(r) + [c["status"], c["motivo"]]
        ws.append(linha)
        fill = FILL_POR_STATUS.get(c["status"], FILL_ABERTO)
        for cel in ws[ws.max_row]:
            cel.fill = fill
    ajustar_largura_colunas(ws)
    formatar_planilha_padrao(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def combinar_workbooks(pares: list) -> bytes:
    """Junta várias planilhas (cada uma já pronta em bytes) num único
    arquivo, uma aba por planilha de origem — usado pelo download
    'Excel combinado' da tela de saúde do cliente. pares: lista de
    (prefixo_nome_aba, xlsx_bytes)."""
    destino = Workbook()
    destino.remove(destino.active)
    nomes_usados = set()
    for prefixo, xlsx_bytes in pares:
        if not xlsx_bytes:
            continue
        origem = load_workbook(io.BytesIO(xlsx_bytes))
        for nome_aba in origem.sheetnames:
            ws_origem = origem[nome_aba]
            nome_final = f"{prefixo} - {nome_aba}"[:31]
            sufixo = 1
            base = nome_final
            while nome_final in nomes_usados:
                sufixo += 1
                nome_final = f"{base[:28]}~{sufixo}"
            nomes_usados.add(nome_final)
            ws_destino = destino.create_sheet(title=nome_final)
            for linha in ws_origem.iter_rows():
                for celula in linha:
                    nova = ws_destino.cell(row=celula.row, column=celula.column, value=celula.value)
                    if celula.has_style:
                        nova.font = celula.font.copy()
                        nova.fill = celula.fill.copy()
                        nova.alignment = celula.alignment.copy()
            ws_destino.freeze_panes = "A2"
            for row_idx in range(1, ws_destino.max_row + 1):
                ws_destino.row_dimensions[row_idx].height = 21
    buf = io.BytesIO()
    destino.save(buf)
    return buf.getvalue()
